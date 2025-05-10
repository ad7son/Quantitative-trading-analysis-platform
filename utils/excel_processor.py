import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
import warnings
import shutil
warnings.filterwarnings('ignore')

def process_excel_file(file_path: str) -> None:
    """
    處理指定的 Excel 檔案，在每個 sheet 的最後加入平均值列，
    並新增一個 sheet 整合所有平均值資料。
    寫入後自動將所有儲存格字型設為微軟正黑體，若無則設為 Arial。

    Parameters
    ----------
    file_path : str
        Excel 檔案的完整路徑
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"找不到檔案：{file_path}")
    
    if not file_path.endswith(('.xlsx', '.xls')):
        raise ValueError("檔案必須是 Excel 檔案（.xlsx 或 .xls）")
    
    try:
        # 嘗試使用不同的引擎讀取檔案
        try:
            # 先嘗試用 openpyxl
            df = pd.read_excel(file_path, engine='openpyxl')
            engine = 'openpyxl'
        except:
            try:
                # 如果失敗，嘗試用 xlrd
                df = pd.read_excel(file_path, engine='xlrd')
                engine = 'xlrd'
            except:
                print(f"警告：無法讀取檔案，請確認檔案格式是否正確")
                return

        # 讀取所有工作表名稱
        excel_file = pd.ExcelFile(file_path, engine=engine)
        sheet_names = excel_file.sheet_names
        
        if len(sheet_names) == 0:
            print(f"警告：檔案沒有工作表")
            return
        
        # 先寫入暫存檔，確保寫入成功再覆蓋原始檔案
        base, ext = os.path.splitext(file_path)
        tmp_path = f"{base}_tmp{ext}"
        all_means = []
        all_mean_sheet_names = []
        with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)
                    if df.empty:
                        continue
                    unnamed_cols = [col for col in df.columns if 'Unnamed' in str(col)]
                    if unnamed_cols:
                        df.set_index(unnamed_cols[0], inplace=True)
                        df.index.name = None
                    numeric_columns = df.select_dtypes(include=['float64', 'int64']).columns
                    if len(numeric_columns) > 0:
                        mean_row = df[numeric_columns].mean()
                        mean_df = pd.DataFrame([mean_row], index=['平均值'])
                        df_with_mean = pd.concat([df, mean_df])
                        df_with_mean.to_excel(writer, sheet_name=sheet_name)
                        all_means.append(mean_row)
                        all_mean_sheet_names.append(sheet_name)
                    else:
                        df.to_excel(writer, sheet_name=sheet_name)
                except Exception as e:
                    continue
            if all_means:
                keep_columns = [
                    '勝率 [%]', '總收益率 [%]', '交易次數', '累計手續費',
                    '平均獲利交易 [%]', '平均虧損交易 [%]', '獲利交易平均持倉時間', '虧損交易平均持倉時間'
                ]
                means_df = pd.concat(all_means, axis=1).T
                means_df.index = all_mean_sheet_names
                means_df = means_df[keep_columns]
                means_df.index.name = None
                means_df = means_df.round(2)
                means_df.to_excel(writer, sheet_name='average')
        try:
            wb = load_workbook(tmp_path)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        try:
                            cell.font = Font(name='Microsoft JhengHei')
                            cell.font = Font(name='Arial')
                        except:
                            cell.font = Font(name='Arial')
                for col in ws.columns:
                    col_letter = col[0].column_letter
                    header = col[0].value
                    if header is not None and str(header).strip() != '':
                        max_length = len(str(header))
                    else:
                        max_length = 3
                    ws.column_dimensions[col_letter].width = max_length * 2 + 2
            wb.save(tmp_path)
        except Exception as e:
            pass
        shutil.move(tmp_path, file_path)

    except Exception as e:
        print(f"處理檔案時發生錯誤：{str(e)}")

if __name__ == "__main__":
    # 使用範例
    file_path = "output/test/TEST_DATA.xlsx"  # 請替換成您要處理的檔案路徑
    try:
        process_excel_file(file_path)
        print("檔案處理完成！")
    except Exception as e:
        print(f"程式執行發生錯誤：{str(e)}") 