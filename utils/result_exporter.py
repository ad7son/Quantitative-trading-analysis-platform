import os
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Dict, Any
from openpyxl import load_workbook

# 設定中文字體
plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei']  # 微軟正黑體
plt.rcParams['axes.unicode_minus'] = False  # 用來正常顯示負號

class ResultExporter:
    """回測結果匯出器
    
    負責將回測結果保存為CSV文件並生成視覺化圖表。
    
    Attributes:
        output_dir (str): 輸出目錄
    """
    
    def __init__(self, output_dir: str = 'output'):
        """
        初始化結果匯出器
        
        Parameters:
        -----------
        output_dir : str
            輸出目錄
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def _translate_column_names(self, df: pd.DataFrame, file_type: str) -> pd.DataFrame:
        """
        將DataFrame的欄位名稱翻譯為中文
        
        Parameters:
        -----------
        df : pd.DataFrame or pd.Series
            要翻譯欄位名稱的DataFrame或Series
        file_type : str
            文件類型，用於選擇對應的翻譯字典
            
        Returns:
        --------
        pd.DataFrame or pd.Series
            欄位名稱已翻譯的DataFrame或Series
        """
        # 統計數據欄位翻譯
        stats_translations = {
            'Start': '開始日期',
            'End': '結束日期',
            'Start Value': '起始資金',
            'End Value': '結束資金',
            'Period': '回測期間',
            'Total Return [%]': '總收益率 [%]',
            'Benchmark Return [%]': '基準回報率 [%]',
            'Max Gross Exposure [%]': '最高總曝險 [%]',
            'Total Fees Paid': '累計手續費',
            'Max Drawdown [%]': '最大回撤 [%]',
            'Max Drawdown Duration': '最大回撤持續時間',
            'Total Trades': '交易次數',
            'Total Closed Trades': '已結束交易次數',
            'Total Open Trades': '未平倉交易次數',
            'Open Trade PnL': '未平倉損益',
            'Buy & Hold Return [%]': '買入持有收益率 [%]',
            'Max. Drawdown [%]': '最大回撤 [%]',
            'Avg. Drawdown [%]': '平均回撤 [%]',
            'Max. Drawdown Duration': '最大回撤持續期',
            'Avg. Drawdown Duration': '平均回撤持續期',
            'Recovery Factor': '恢復因子',
            'Sharpe Ratio': '夏普比率',
            'Sortino Ratio': '索提諾比率',
            'Calmar Ratio': '卡瑪比率',
            'Omega Ratio': '歐米伽比率',
            'Value at Risk': '風險價值',
            'Expected Shortfall': '預期損失',
            'Avg. Trade [%]': '平均交易收益 [%]',
            'Avg. Win [%]': '平均盈利 [%]',
            'Avg. Loss [%]': '平均虧損 [%]',
            'Win Rate [%]': '勝率 [%]',
            'Best Trade [%]': '最佳交易 [%]',
            'Worst Trade [%]': '最差交易 [%]',
            'Avg Winning Trade [%]': '平均獲利交易 [%]',
            'Avg Losing Trade [%]': '平均虧損交易 [%]',
            'Avg Winning Trade Duration': '獲利交易平均持倉時間',
            'Avg Losing Trade Duration': '虧損交易平均持倉時間',
            'Avg. Holding Time': '平均持倉時間',
            'Trades': '交易次數',
            'Wins': '盈利次數',
            'Losses': '虧損次數',
            'Win/Loss Ratio': '盈虧比',
            'Profit Factor': '獲利因子',
            'Expectancy': '期望值',
            'SqN': '系統品質數',
            'K-factor': 'K因子',
            'Risk-Return Ratio': '風險收益比',
            'Daily Sharpe': '日夏普比率',
            'Daily Sortino': '日索提諾比率',
            'Daily Calmar': '日卡瑪比率',
            'Daily Omega': '日歐米伽比率',
            'Daily Value at Risk': '日風險價值',
            'Daily Expected Shortfall': '日預期損失',
            'Daily Avg. Trade [%]': '日平均交易收益 [%]',
            'Daily Avg. Win [%]': '日平均盈利 [%]',
            'Daily Avg. Loss [%]': '日平均虧損 [%]',
            'Daily Win Rate [%]': '日勝率 [%]',
            'Daily Best Trade [%]': '日最佳交易 [%]',
            'Daily Worst Trade [%]': '日最差交易 [%]',
            'Daily Avg. Holding Time': '日平均持倉時間',
            'Daily Trades': '日交易次數',
            'Daily Wins': '日盈利次數',
            'Daily Losses': '日虧損次數',
            'Daily Win/Loss Ratio': '日盈虧比',
            'Daily Profit Factor': '日獲利因子',
            'Daily Expectancy': '日期望值',
            'Daily SqN': '日系統品質數',
            'Daily K-factor': '日K因子',
            'Daily Risk-Return Ratio': '日風險收益比'
        }
        
        # 交易記錄欄位翻譯
        trades_translations = {
            'Exit Trade Id': '出場交易ID',
            'Column': '欄位',
            'Size': '交易數量',
            'Entry Timestamp': '進場時間',
            'Avg Entry Price': '平均進場價格',
            'Entry Fees': '進場費用',
            'Exit Timestamp': '出場時間',
            'Avg Exit Price': '平均出場價格',
            'Exit Fees': '出場費用',
            'PnL': '盈虧',
            'Return': '收益率',
            'Direction': '方向',
            'Status': '狀態',
            'Position Id': '持倉ID'
        }
        
        # 持倉記錄欄位翻譯
        positions_translations = {
            'Position Id': '持倉ID',
            'Column': '欄位',
            'Size': '持倉數量',
            'Entry Timestamp': '進場時間',
            'Avg Entry Price': '平均進場價格',
            'Entry Fees': '進場費用',
            'Exit Timestamp': '出場時間',
            'Avg Exit Price': '平均出場價格',
            'Exit Fees': '出場費用',
            'PnL': '盈虧',
            'Return': '收益率',
            'Direction': '方向',
            'Status': '狀態'
        }
        
        # 選擇對應的翻譯字典
        translations = {
            'stats': stats_translations,
            'trades': trades_translations,
            'positions': positions_translations
        }
        
        # 翻譯欄位名稱
        if file_type in translations:
            if isinstance(df, pd.Series):
                # 如果是Series，直接使用rename
                df = df.rename(translations[file_type])
            else:
                # 如果是DataFrame，使用rename(columns=...)
                df = df.rename(columns=translations[file_type])
        
        return df
    
    def save_statistics(self, results: Dict[str, Any]) -> None:
        """
        保存統計數據到CSV文件
        
        Parameters:
        -----------
        results : dict
            回測結果字典
        """
        # 保存基本統計數據
        stats_df = self._translate_column_names(results['stats'], 'stats')
        # 將 Series 轉換為 DataFrame 並設置索引名稱
        stats_df = pd.DataFrame(stats_df)
        stats_df.index.name = '指標'
        stats_df.to_csv(f'{self.output_dir}/詳細數據.csv')
        
        # 保存交易記錄
        trades_df = self._translate_column_names(results['trades'], 'trades')
        trades_df.to_csv(f'{self.output_dir}/交易紀錄.csv', index=False)
        
        # 保存每日收益
        daily_returns = results['daily_returns']
        daily_returns.name = '每單位收益率'
        daily_returns.to_csv(f'{self.output_dir}/單位收益.csv')
        
        # 保存持倉記錄
        positions_df = self._translate_column_names(results['positions'], 'positions')
        positions_df.to_csv(f'{self.output_dir}/持倉記錄.csv', index=False)

        return stats_df
            
    def plot_results(self, data: pd.DataFrame, results: Dict[str, Any], 
                    all_stats: Dict[str, Any], strategy_params) -> None:
        """
        繪製回測結果圖表
        
        Parameters:
        -----------
        data : pd.DataFrame
            原始數據
        results : dict
            回測結果
        all_stats : dict
            所有統計數據
        strategy_params
            策略參數
        """
        # 價格和移動平均線圖
        plt.figure(figsize=(20, 10))
        data['close'].plot(label='收盤價', color='black')

        for p in strategy_params.short_periods:
            data['close'].ewm(span=p, adjust=False).mean().plot(label=f'Short {p}', linewidth=0.5, color='blue')
        for p in strategy_params.long_periods:
            data['close'].ewm(span=p, adjust=False).mean().plot(label=f'Long {p}', linewidth=0.5, color='red')

        plt.title('價格和移動平均線')
        plt.legend()
        plt.savefig(f'{self.output_dir}/價格和移動平均線.png')
        plt.close()
        
        # # 價格和移動平均線圖
        # plt.figure(figsize=(15, 6))
        # # data['close'].rolling(window=strategy_params.period).mean().plot(label='5日 SMA')
        # data['close'].plot(label='收盤價', color='black')
        # short_ma = data['close'].rolling(window=strategy_params.short_period).mean()
        # mid_ma = data['close'].rolling(window=strategy_params.mid_period).mean()
        # long_ma = data['close'].rolling(window=strategy_params.long_period).mean()
        # short_ma.plot(label=f'{strategy_params.short_period}日 短線', color='blue')
        # mid_ma.plot(label=f'{strategy_params.mid_period}日 中線', color='green')
        # long_ma.plot(label=f'{strategy_params.long_period}日 慢線', color='red')
        # plt.title('價格和移動平均線')
        # plt.legend()
        # plt.savefig(f'{self.output_dir}/價格和移動平均線.png')
        # plt.close()

        # 投資組合價值圖
        plt.figure(figsize=(15, 6))
        results['portfolio'].value().plot()
        plt.title('投資組合價值')
        plt.savefig(f'{self.output_dir}/投資組合價值.png')
        plt.close()
        
        # 每月度收益熱力圖
        plt.figure(figsize=(15, 8))
        daily_returns = all_stats['daily_returns']
        daily_returns_matrix = daily_returns.to_frame()
        daily_returns_matrix.index = pd.to_datetime(daily_returns_matrix.index)
        daily_returns_matrix['month'] = daily_returns_matrix.index.month
        daily_returns_matrix['year'] = daily_returns_matrix.index.year
        pivot_table = daily_returns_matrix.pivot_table(
            values=daily_returns_matrix.columns[0],
            index='year',
            columns='month',
            aggfunc='sum'
        )
        sns.heatmap(pivot_table, annot=True, fmt='.2%', cmap='RdYlGn', center=0)
        plt.title('月度收益熱力圖')
        plt.savefig(f'{self.output_dir}/月度收益熱力圖.png')
        plt.close()
        
        # 權益曲線和回撤圖
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(15, 10))
        
        # 權益曲線
        portfolio_value = results['portfolio'].value()
        portfolio_value.plot(ax=ax1)
        ax1.set_title('權益曲線')
        
        # 回撤圖
        drawdown = results['portfolio'].drawdown()
        drawdown.plot(ax=ax2, color='red')
        ax2.set_title('回撤')
        
        plt.tight_layout()
        plt.savefig(f'{self.output_dir}/權益曲線和回撤圖.png')
        plt.close()

    def export_all(self, data: pd.DataFrame, results: Dict[str, Any], 
                  all_stats: Dict[str, Any], strategy_params) -> pd.DataFrame:
        """
        匯出個別股票的所有結果
        
        Parameters:
        -----------
        data : pd.DataFrame
            原始數據
        results : dict
            回測結果
        all_stats : dict
            所有統計數據
        strategy_params
            策略參數
        """
        stats_df = self.save_statistics(all_stats)
        self.plot_results(data, results, all_stats, strategy_params)

        # 從輸出路徑中提取股票代碼
        stock_id = self.output_dir.split('/')[-3]  # output/SMAC/2330/1D5D/ -> 2330
        
        # 將 DataFrame 的列名改為股票代碼
        stats_df.columns = [stock_id]

        return stats_df 

    def write_df_to_excel(self, file_path: str, sheet_name: str, df: pd.DataFrame):
        """
        將 DataFrame 寫入 Excel 文件
        
        Parameters:
        -----------
        file_path : str
            Excel 文件路徑
        sheet_name : str
            工作表名稱
        df : pd.DataFrame
            要寫入的數據框
        """
        # 確保輸出目錄存在
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # 轉置 DataFrame，使指標成為列名，數據成為行
        df_transposed = df.transpose()
        
        # 檢查文件是否存在且是有效的 Excel 文件
        is_valid_excel = False
        if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
            try:
                # 嘗試讀取文件的前幾個字節來檢查是否為有效的 Excel 文件
                with open(file_path, 'rb') as f:
                    header = f.read(4)
                    # Excel 文件的魔數是 PK\x03\x04
                    is_valid_excel = header.startswith(b'PK\x03\x04')
            except Exception:
                is_valid_excel = False
        
        try:
            if is_valid_excel:
                try:
                    # 嘗試讀取現有的 Excel 文件
                    book = load_workbook(file_path)
                    
                    # 檢查工作表是否存在
                    if sheet_name in book.sheetnames:
                        try:
                            # 如果工作表存在，讀取現有數據
                            existing_df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)
                            
                            # 將新數據追加到現有數據後面
                            combined_df = pd.concat([existing_df, df_transposed], axis=0)
                            
                            # 寫入合併後的數據
                            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                # 寫入合併後的工作表
                                combined_df.to_excel(writer, sheet_name=sheet_name)
                        except Exception as e:
                            print(f"讀取現有工作表時發生錯誤: {e}")
                            # 如果讀取失敗，直接覆蓋寫入
                            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                df_transposed.to_excel(writer, sheet_name=sheet_name)
                    else:
                        # 如果工作表不存在，直接追加新工作表
                        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                            df_transposed.to_excel(writer, sheet_name=sheet_name)
                except Exception as e:
                    print(f"讀取 Excel 文件時發生錯誤: {e}")
                    # 如果文件損壞，刪除並重新創建
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        print(f"刪除損壞的文件時發生錯誤: {e}")
                    
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df_transposed.to_excel(writer, sheet_name=sheet_name)
            else:
                # 如果文件不存在或不是有效的 Excel 文件，則創建新文件
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        print(f"刪除無效的文件時發生錯誤: {e}")
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df_transposed.to_excel(writer, sheet_name=sheet_name)
        except Exception as e:
            print(f"寫入 Excel 文件時發生錯誤: {e}")
            # 最後的嘗試：直接覆蓋寫入
            try:
                # 確保文件不存在
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        print(f"刪除文件時發生錯誤: {e}")
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df_transposed.to_excel(writer, sheet_name=sheet_name)
            except Exception as e:
                print(f"最終嘗試寫入 Excel 文件時發生錯誤: {e}")
                raise

        
